#!/usr/bin/env python3
"""Inspect the Active Workers Directory xlsx to understand its schema."""

import sys

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

XLSX_PATH = "app/config/VRTX_-_Active_Workers_Directory.xlsx"

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

print("=" * 60)
print(f"FILE: {XLSX_PATH}")
print(f"SHEET NAMES: {wb.sheetnames}")
print("=" * 60)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    print(f"\n{'=' * 60}")
    print(f"SHEET: '{sheet_name}'")
    print(f"TOTAL ROWS (including header): {len(rows)}")

    if len(rows) == 0:
        print("  (empty sheet)")
        continue

    # Headers
    headers = list(rows[0])
    print(f"DATA ROWS: {len(rows) - 1}")
    print(f"\nCOLUMN HEADERS ({len(headers)} columns):")
    for i, h in enumerate(headers):
        print(f"  [{i}] {h!r}")

    # First 3 data rows as dicts
    print(f"\nFIRST 3 DATA ROWS:")
    for row_idx, row in enumerate(rows[1:4], start=2):
        print(f"\n  --- Row {row_idx} ---")
        for h, v in zip(headers, row):
            print(f"    {h}: {v!r}")

    # Find email-like columns
    email_cols = []
    for i, h in enumerate(headers):
        if h and "email" in str(h).lower():
            email_cols.append((i, h))
        elif h and "mail" in str(h).lower():
            email_cols.append((i, h))
        elif h and "@" in str(h).lower():
            email_cols.append((i, h))

    if email_cols:
        print(f"\nEMAIL-LIKE COLUMNS:")
        for col_idx, col_name in email_cols:
            sample_vals = [rows[r][col_idx] for r in range(1, min(8, len(rows)))]
            print(f"  Column '{col_name}' (idx {col_idx}):")
            for sv in sample_vals:
                print(f"    {sv!r}")
    else:
        print("\nNO EMAIL-LIKE COLUMN FOUND. Showing all columns with sample values:")
        for i, h in enumerate(headers):
            sample_vals = [rows[r][i] for r in range(1, min(4, len(rows)))]
            print(f"  [{i}] '{h}': {sample_vals}")

    # Check for columns that might contain email addresses by content
    print(f"\nSCANNING FOR '@' IN CELL VALUES (first 10 rows):")
    for i, h in enumerate(headers):
        for r in range(1, min(11, len(rows))):
            val = rows[r][i]
            if val and isinstance(val, str) and "@" in val:
                print(f"  Column [{i}] '{h}' row {r + 1}: {val!r}")
                break

wb.close()
print("\n" + "=" * 60)
print("DONE")
